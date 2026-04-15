import os
import pandas as pd
from openpyxl import load_workbook
import shutil


# =====================================基础配置===============================

# 数据导入
total_path = r'C:\Users\Yuyu\Desktop\收款\汇总表.xlsx'
bank_path = r'C:\Users\Yuyu\Desktop\收款\银行流水.xlsx'
customer = pd.read_excel(total_path,sheet_name='客户代码')
total = pd.read_excel(total_path,sheet_name='汇总数据')
bank = pd.read_excel(bank_path,sheet_name='SHEET1')
output_dir = r'C:\Users\Yuyu\Desktop\收款\凭证导入'
os.makedirs(output_dir, exist_ok=True)

# =======================================结束==================================



# =============================银行数据清洗生成凭证明细表==========================

# 正文修正 筛选去除表头银行数据汇总等不必要行
bank = bank.iloc[12:]
# 筛选收款发生额不为0的数据（排除付款的流水项目）
bank = bank[bank.iloc[:, 10].notna()]
# 筛选付款账户名称为公司的项目（排除个人收款）
bank = bank[bank.iloc[:, 5].astype(str).str.contains('公司', na=False)]
# 银行流水表里只提取需要数据 按位置提取 A (0), F (5), K (10) 三列数据
bank = bank.iloc[:, [0, 5, 10]]
# 重新构建标题行 方便匹配客户代码
bank.columns = ['交易日期', '客户名称', '收款发生额']
# 将客户名称匹配上对应SAP系统里的客户代码（类似xlookup）
bank_selected = pd.merge(
                bank,
                customer,
                on='客户名称',  # 按按户名称匹配
                how='left'  # 左连接，保留所有记录
 )
# 转换格式 修正客户代码浮点转字符串 收款转换为浮点数格式
bank_selected['客户代码']=bank_selected['客户代码'].astype(int).astype(str)
bank_selected['收款发生额']=bank_selected['收款发生额'].astype(float)
# 增加列项 设置需要列便于后续生成凭证
bank_selected['回款性质']="试剂款"
bank_selected['现金流指定']="A01"
bank_selected['银行代码']="10021031"
bank_selected['特别总账标志']="A"
bank_selected['凭证号']="A"
bank_selected['凭证文本']=bank_selected['客户代码'] + "/"+bank_selected['客户名称'] + "/"+bank_selected['回款性质']
# 构建行标题
column_order = [
    '交易日期', '客户代码', '客户名称', '收款发生额',
    '回款性质', '银行代码', '凭证文本',
    '特别总账标志', '现金流指定'
]
# 标题修正
bank_selected = bank_selected[column_order]
print(f"凭证明细表 {bank_selected}")

# =================================结束===================================



# ==================================按模板批量生成凭证========================

# 删除凭证导入文件夹里的文件（避免后续重复导入）
shutil.rmtree(output_dir, ignore_errors=True)
os.makedirs(output_dir, exist_ok=True)
# 读取模板文件数据
template_path = r'C:\Users\Yuyu\Desktop\收款\汇总表.xlsx'
# 读取凭证明细表数据
Batch_voucher = bank_selected
# 遍历每一行数据 按照模板批量生成凭证
for index, row in Batch_voucher.iterrows():
# 读取模版工作表
    wb = load_workbook(template_path)
    ws = wb['模板']
    ws.title = "Sheet1"
# 指定要保留的工作表名称
    target_sheet_name = 'Sheet1'
# 获取当前所有工作表名称（复制一份列表，避免遍历时修改）
    all_sheets = wb.sheetnames[:]
# 删除除目标工作表以外的所有工作表（因为汇总表这个表里还有别的sheet表 不需要复制过去）
    for sheet_name in all_sheets:
        if sheet_name != target_sheet_name:
            wb.remove(wb[sheet_name])
# wb 中只剩 '模板' 这一个工作表
    ws = wb[target_sheet_name]
# 处理凭证日期过账日期 将格式改成为8位字符串
    voucher_date = str(row['交易日期'])
    posting_date = voucher_date

# ------------------------数据填充部分---------------------------------
    # 填充抬头部分（第三行）
    ws['B3'] = voucher_date  # 凭证日期
    ws['C3'] = posting_date  # 过账日期
    ws['E3'] = '收款'  # 参照
    ws['F3'] = row['客户代码']  # 公司名称
    ws['G3'] = 'DA'  # 凭证类型
    ws['H3'] = 1030  # 公司代码
    ws['I3'] = 'CNY'  # 货币

    # 第一个行项目（第四行）
    ws['B4'] = 40
    ws['C4'] = row['银行代码']
    ws['E4'] = row['收款发生额']
    ws['S4'] = row['凭证文本']
    ws['T4'] = row['现金流指定']

    # 第二个行项目（第五行）
    ws['B5'] = 19
    ws['C5'] = row['客户代码']
    ws['D5'] = row['特别总账标志']
    ws['E5'] = row['收款发生额']
    ws['S5'] = row['凭证文本']

    # 生成文件名（替换特殊字符）
    client_code = str(row['客户代码'])
    client_name = row['客户名称'].replace('/', '_').replace('\\', '_')
    filename = f"{client_code}_{client_name}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # 保存文件
    wb.save(filepath)
# -------------------------------------------------结束-----------------------------------

# 读取原有的汇总数据 将提取的数据追加到汇总表
original_total = pd.read_excel(total_path, sheet_name='汇总数据')
# 拼接新旧数据
new_total = pd.concat([original_total, bank_selected], ignore_index=True)
# 写回 Excel（覆盖 sheet）
with pd.ExcelWriter(total_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    new_total.to_excel(writer, sheet_name='汇总数据', index=False)
print(bank_selected)

# ==========================================结束===========================================

# ============================SAP操作项目===================================