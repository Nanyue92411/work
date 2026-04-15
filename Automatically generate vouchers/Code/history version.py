import subprocess
import time

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import ddddocr
from datetime import datetime, timedelta
import os
import glob
import shutil
from selenium.webdriver.support.ui import WebDriverWait
import undetected_chromedriver as uc
#========================基础配置==============================

URL = "https://corp.bank.ecitic.com/cotb/electronic/login.html"#网址
DOWNLOAD_folder = r"C:\Users\DELL\Desktop" #下载目录
TARGET_folder = r"D:\银行流水\2026年4月"  #目标目录

# ========================结束==================================


# ===========================网页操作=================================

# 关闭所有 Chrome 进程 防止其他进程干扰 （可选，会丢失已有会话）
subprocess.run('taskkill /f /im chrome.exe', shell=True, capture_output=True)
time.sleep(2)
# 等待浏览器完全打开
chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
user_data_dir = r"D:\selenium_profile"
cmd = f'"{chrome_path}" --remote-debugging-port=9222 --user-data-dir="{user_data_dir}"'
subprocess.Popen(cmd, shell=True)
# time.sleep(3)

# 连接到调试浏览器
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
driver = uc.Chrome()
# 打开登录页面
driver.get(URL)
print("当前页面标题:", driver.title)
time.sleep(3)
# 自动填写用户名
Username = driver.find_element(By.NAME, "userCode")
Username.send_keys("账号")
# 等待用户输入密码
time.sleep(15)

# def get_captcha_text(browser, img_selector):
#     ocr = ddddocr.DdddOcr(show_ad=False)
#     captcha_el = browser.find_element(By.CSS_SELECTOR, img_selector)
#     try:
#         img_bytes = captcha_el.screenshot_as_png
#         result = ocr.classification(img_bytes)
#         print(f"验证码识别结果：{result}")
#         return result
#     except Exception as e:
#         print(f"截图失败：{e}，改用下载方式")
#         img_url = captcha_el.get_attribute("src")
#         cookies = {c["name"]: c["value"] for c in browser.get_cookies()}
#         img_bytes = requests.get(img_url, cookies=cookies).content
#         result = ocr.classification(img_bytes)
#         print(f"验证码识别结果：{result}")
#         return result
#
# MAX_TRIES = 15
#
# for attempt in range(1, MAX_TRIES + 1):
#     print(f"第 {attempt} 次尝试登录...")
#
#     # ── 自动识别验证码 ──
#     # img_selector 换成 F12 看到的验证码 img 标签的选择器
#     captcha_text = get_captcha_text(driver, "img#verifyCode")
#
#     captcha_input = driver.find_element(By.ID, "verifyCodeInput")
#     captcha_input.clear()
#     captcha_input.send_keys(captcha_text)
#
#     login_btn = driver.find_element(By.XPATH, "//button[@data-i18n-key='i0008']")
#     login_btn.click()
#
#     try:
#         WebDriverWait(driver, 5).until(
#             EC.presence_of_element_located((By.XPATH, "//a[@class='first-menu']"))
#         )
#         print("登录成功！")
#         break
#     except:
#         try:
#             close_btn = WebDriverWait(driver, 3).until(
#                 EC.element_to_be_clickable((By.XPATH, "//button[text()='关闭']"))
#             )
#             close_btn.click()
#             print("验证码错误，关闭弹窗，重试...")
#             time.sleep(1)
#         except:
#             print("未找到关闭按钮，停止重试")
#             break
# else:
#     print(f"达到最大尝试次数 {MAX_TRIES}，登录失败。")


# ========================================================================
# 处理验证码（自动识别并填入）
# 1. 定位验证码图片元素并截图保存
captcha_element = driver.find_element(By.ID, "verifyCode")
# 验证码截图
captcha_element.screenshot("captcha.png")
# 2. 创建识别器对象
ocr = ddddocr.DdddOcr()
# 3. 读取图片 (二进制模式)
with open('captcha.png', 'rb') as f:
    img_bytes = f.read()
# 4. 识别并打印结果
res = ocr.classification(img_bytes)
print('识别出的验证码为: ' + res)
# 5. 输入验证码
driver.find_element(By.ID, "verifyCodeInput").send_keys(res)
# 点击登录按钮
login_btn = driver.find_element(By.XPATH, "//button[@data-i18n-key='i0008']")
login_btn.click()
print("已点击登录按钮")

# 验证码输入错误弹出窗
close_btn = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//button[text()='关闭']"))
)
close_btn.click()

MAX_TRIES = 15

for attempt in range(1, MAX_TRIES + 1):
    print(f"第 {attempt} 次尝试登录...")
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//a[@class='first-menu']"))
        )
        print("登录成功！")
        break
    except:
        try:
            close_btn = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button[text()='关闭']"))
            )
            close_btn.click()
            print("验证码错误，关闭弹窗，重试...")
            time.sleep(1)
        except:
            print("未找到关闭按钮，停止重试")
            break
else:
    print(f"达到最大尝试次数 {MAX_TRIES}，登录失败。")



# ========================账户余额及交易明细查询=============
# 点击账户明细
menu = driver.find_element(By.XPATH, "(//a[@class='first-menu'])[2]")
menu.click()
# 选择账号
# 1. 点击只读输入框，触发下拉面板
input_box = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.ID, "bsnAcc"))
)
input_box.click()
# 2. 等待下拉选项出现 选择网银账户 option = driver.find_element(By.XPATH, "//li[@data-value='123456789']")
option = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//li[contains(text(), '账号')]"))
)
option.click()

# 处理日期选择器 选择昨日银行流水
# 第一步：算出目标日期
today     = datetime.now()
yesterday = today - timedelta(days=1)
# 日历的 data-ymd 格式是 2026-4-13（月和日没有补零）
# 所以格式化时用 %-m 和 %-d（Linux），Windows 用 %#m 和 %#d
date_str = f"{yesterday.year}-{yesterday.month}-{yesterday.day}"
print(date_str)  # 2026-4-13

# 第二步：点击输入框弹出日历
start_input = driver.find_element(By.NAME, "startDate")
start_input.click()

# 第三步：处理开始日期
wait = WebDriverWait(driver, 10)
wait.until(EC.visibility_of_element_located((By.ID, "jedatebox")))

# 第四步：直接找到目标日期的 li 并点击
target = driver.find_element(By.XPATH, f"//li[@data-ymd='{date_str}']")
target.click()

# # 第五步：点确定
# start_input = driver.find_element(By.NAME, "endDate")
# start_input.click()

# 处理截止日期
wait = WebDriverWait(driver, 10)
wait.until(EC.visibility_of_element_located((By.ID, "jedatebox")))

# 直接找到目标日期的 li 并点击
target = driver.find_element(By.XPATH, f"//li[@data-ymd='{date_str}']")
target.click()

# 点击查询按钮
qry = driver.find_element(By.XPATH, "//button[@ng-click='qry()']")
qry.click()

# 点击下载
download = driver.find_element(By.XPATH, "//a[text()='全部下载']")
download.click()
# 选择Excel下载
download_select = driver.find_element(By.XPATH, "//a[text()='Excel下载']")
download_select.click()

# 确保目标文件夹存在
os.makedirs(TARGET_folder, exist_ok=True)
# 获取原文件路径
old_path = glob.glob(os.path.join(DOWNLOAD_folder, "COBP*.xlsx"))[0]

# 复制到目标文件夹并重命名为日期.xlsx
new_path = os.path.join(TARGET_folder, f"{date_str}.xlsx")
shutil.move(old_path, new_path)


# =================================================================================================
# ========================================下面为银行流水数据处理=====================================
import os
import pandas as pd
from openpyxl import load_workbook
import shutil


# =====================================基础配置===============================

# 数据导入
total_path = r'C:\Users\Yuyu\Desktop\收款\汇总表.xlsx'
bank_path = rf'C:\Users\Yuyu\Desktop\收款\{date_str}.xlsx'
customer = pd.read_excel(total_path,sheet_name='客户代码')
total = pd.read_excel(total_path,sheet_name='汇总数据')
bank = pd.read_excel(bank_path,sheet_name='SHEET1')
output_dir = r'C:\Users\Yuyu\Desktop\收款\凭证导入'
os.makedirs(output_dir, exist_ok=True)

# =======================================结束==================================



# =============================银行数据清洗生成凭证明细表==========================

# 正文修正 筛选去除表头银行数据汇总等不必要行
bank = bank.iloc[12:]
# 筛选收款发生额不为0的数据（排除付款的流水项目）
bank = bank[bank.iloc[:, 10].notna()]
# 筛选付款账户名称为公司的项目（排除个人收款）
bank = bank[bank.iloc[:, 5].astype(str).str.contains('公司', na=False)]
# 银行流水表里只提取需要数据 按位置提取 A (0), F (5), K (10) 三列数据
bank = bank.iloc[:, [0, 5, 10]]
# 重新构建标题行 方便匹配客户代码
bank.columns = ['交易日期', '客户名称', '收款发生额']
# 将客户名称匹配上对应SAP系统里的客户代码（类似xlookup）
bank_selected = pd.merge(
                bank,
                customer,
                on='客户名称',  # 按按户名称匹配
                how='left'  # 左连接，保留所有记录
 )
# 转换格式 修正客户代码浮点转字符串 收款转换为浮点数格式
bank_selected['客户代码']=bank_selected['客户代码'].astype(int).astype(str)
bank_selected['收款发生额']=bank_selected['收款发生额'].astype(float)
# 增加列项 设置需要列便于后续生成凭证
bank_selected['回款性质']="试剂款"
bank_selected['现金流指定']="A01"
bank_selected['银行代码']="10021031"
bank_selected['特别总账标志']="A"
bank_selected['凭证号']=""
bank_selected['凭证文本']=bank_selected['客户代码'] + "/"+bank_selected['客户名称'] + "/"+bank_selected['回款性质']
# 构建行标题
column_order = [
    '交易日期', '客户代码', '客户名称', '收款发生额',
    '回款性质', '银行代码', '凭证文本',
    '特别总账标志', '现金流指定'
]
# 标题修正
bank_selected = bank_selected[column_order]
print(f"凭证明细表 {bank_selected}")

# =================================结束===================================



# ==================================按模板批量生成凭证========================

# 删除凭证导入文件夹里的文件（避免后续重复导入）
shutil.rmtree(output_dir, ignore_errors=True)
os.makedirs(output_dir, exist_ok=True)
# 读取模板文件数据
template_path = r'C:\Users\Yuyu\Desktop\收款\汇总表.xlsx'
# 读取凭证明细表数据
Batch_voucher = bank_selected
# 遍历每一行数据 按照模板批量生成凭证
for i, (index, row) in enumerate(Batch_voucher.iterrows(), start=1):
# 读取模版工作表
    wb = load_workbook(template_path)
    ws = wb['模板']
    ws.title = "Sheet1"
# 指定要保留的工作表名称
    target_sheet_name = 'Sheet1'
# 获取当前所有工作表名称（复制一份列表，避免遍历时修改）
    all_sheets = wb.sheetnames[:]
# 删除除目标工作表以外的所有工作表（因为汇总表这个表里还有别的sheet表 不需要复制过去）
    for sheet_name in all_sheets:
        if sheet_name != target_sheet_name:
            wb.remove(wb[sheet_name])
# wb 中只剩 '模板' 这一个工作表
    ws = wb[target_sheet_name]
# 处理凭证日期过账日期 将格式改成为8位字符串
    voucher_date = str(row['交易日期'])
    posting_date = voucher_date

# ------------------------数据填充部分---------------------------------
    # 填充抬头部分（第三行）
    ws['B3'] = voucher_date  # 凭证日期
    ws['C3'] = posting_date  # 过账日期
    ws['E3'] = '收款'  # 参照
    ws['F3'] = row['客户代码']  # 公司名称
    ws['G3'] = 'DA'  # 凭证类型
    ws['H3'] = 1030  # 公司代码
    ws['I3'] = 'CNY'  # 货币

    # 第一个行项目（第四行）
    ws['B4'] = 40
    ws['C4'] = row['银行代码']
    ws['E4'] = row['收款发生额']
    ws['S4'] = row['凭证文本']
    ws['T4'] = row['现金流指定']

    # 第二个行项目（第五行）
    ws['B5'] = 19
    ws['C5'] = row['客户代码']
    ws['D5'] = row['特别总账标志']
    ws['E5'] = row['收款发生额']
    ws['S5'] = row['凭证文本']

    # 生成文件名（替换特殊字符）
    client_code = str(row['客户代码'])
    client_name = row['客户名称'].replace('/', '_').replace('\\', '_')
    filename = f"{i:03d}_{client_code}_{client_name}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # 保存文件
    wb.save(filepath)
# -------------------------------------------------结束-----------------------------------

# 读取原有的汇总数据 将提取的数据追加到汇总表
original_total = pd.read_excel(total_path, sheet_name='汇总数据')
# 拼接新旧数据
new_total = pd.concat([original_total, bank_selected], ignore_index=True)
# 写回 Excel（覆盖 sheet）
with pd.ExcelWriter(total_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    new_total.to_excel(writer, sheet_name='汇总数据', index=False)
print(bank_selected)

# ==========================================结束===========================================

